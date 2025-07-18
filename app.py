import os
from fastapi import FastAPI, File, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from tempfile import NamedTemporaryFile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from starlette.middleware.cors import CORSMiddleware

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/", response_class=HTMLResponse)
async def get_home():
    with open("static/index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.post("/process")
async def process_excel(
    input_file: UploadFile = File(...),
    output_file: UploadFile = File(...)
):
    compare_columns = ['intellimindz(SERP)', 'Proxy  (SERP)', 'online job support(SERP)']
    arrow_colors = {"↑": "008000", "↓": "FF0000", "↓↓": "FF0000", "→": "808080"}

    input_path = f"/tmp/temp_{input_file.filename}"
    output_path = f"/tmp/temp_{output_file.filename}"

    with open(input_path, "wb") as f:
        f.write(await input_file.read())
    with open(output_path, "wb") as f:
        f.write(await output_file.read())

    result_file = "/tmp/SERP_Comparison_Result.xlsx"
    with pd.ExcelWriter(result_file, engine='openpyxl') as writer:
        sheet_names = pd.ExcelFile(input_path).sheet_names
        for sheet in sheet_names:
            try:
                df_week2 = pd.read_excel(input_path, sheet_name=sheet)
                df_week3 = pd.read_excel(output_path, sheet_name=sheet)
                comparison_df = df_week3[['S.No', 'Technology']].copy()
                difference_df = df_week3[['S.No', 'Technology']].copy()
                for col in compare_columns:
                    if col not in df_week2.columns or col not in df_week3.columns:
                        continue
                    results, diffs = [], []
                    for prev, curr in zip(df_week2[col], df_week3[col]):
                        if pd.isna(prev) or pd.isna(curr):
                            results.append("→ NA")
                            diffs.append("NA")
                            continue
                        if prev == 0 and curr == 0:
                            results.append("↓↓ 0")
                            diffs.append(0)
                        elif curr < prev:
                            results.append(f"↑ {curr}")
                            diffs.append(abs(curr - prev))
                        elif curr > prev:
                            results.append(f"↓ {curr}")
                            diffs.append(abs(curr - prev))
                        else:
                            results.append(f"→ {curr}")
                            diffs.append(0)
                    comparison_df[col] = results
                    difference_df[col] = diffs
                comparison_df.to_excel(writer, sheet_name=f"{sheet}_Comparison", index=False)
                difference_df.to_excel(writer, sheet_name=f"{sheet}_Difference", index=False)
            except Exception as e:
                print(f"Error in sheet {sheet}: {e}")

    # Load workbook and apply formatting
    wb = load_workbook(result_file)

    # Define fills
    fill_green = PatternFill(start_color="00EA00", end_color="00EA00", fill_type="solid")
    fill_red = PatternFill(start_color="FF2121", end_color="FF2121", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    comparison_legend = [
        ["Symbol:", "Meaning:"],
        ["↑", "Current rank is better (improved)."],
        ["↓", "Current rank is worse (declined)."],
        ["↓↓ 0", "Both weeks had default value 0 (no data)."],
        ["→", "No change in rank."],
        ["→ NA", "Data missing in one or both weeks."]
    ]

    difference_legend = [
        ["Color:", "Meaning:"],
        ["Green", "Rank improved."],
        ["Red", "Rank declined."],
        ["Yellow", "No change (difference is 0)."],
        ["White", "Default zero value in both weeks."]
    ]

    for sheet in wb.sheetnames:
        if sheet.endswith("_Comparison"):
            ws_cmp = wb[sheet]
            # Write comparison legend starting at column H (8)
            for r_idx, row in enumerate(comparison_legend, 1):
                for c_idx, val in enumerate(row, 0):  # zero-based offset for column H
                    ws_cmp.cell(row=r_idx, column=8 + c_idx).value = val

            sheet_base = sheet.replace("_Comparison", "")
            diff_sheet = f"{sheet_base}_Difference"
            if diff_sheet not in wb.sheetnames:
                continue
            ws_diff = wb[diff_sheet]

            # Write difference legend starting at column J (10)
            for r_idx, row in enumerate(difference_legend, 1):
                for c_idx, val in enumerate(row, 0):
                    ws_diff.cell(row=r_idx, column=10 + c_idx).value = val

            for col in range(3, ws_cmp.max_column + 1):
                for row in range(2, ws_cmp.max_row + 1):
                    cell_cmp = ws_cmp.cell(row=row, column=col)
                    cell_diff = ws_diff.cell(row=row, column=col)
                    value = str(cell_cmp.value)

                    if value.startswith("↓↓"):
                        cell_cmp.font = Font(color=arrow_colors["↓↓"], bold=True)
                        if value.strip() == "↓↓ 0":
                            cell_diff.fill = fill_white
                        else:
                            cell_diff.fill = fill_red
                    elif value.startswith("↑"):
                        cell_cmp.font = Font(color=arrow_colors["↑"], bold=True)
                        cell_diff.fill = fill_green
                    elif value.startswith("↓"):
                        cell_cmp.font = Font(color=arrow_colors["↓"], bold=True)
                        cell_diff.fill = fill_red
                    elif value.startswith("→"):
                        cell_cmp.font = Font(color=arrow_colors["→"], bold=True)

                    if cell_diff.value == 0 and value.strip() != "↓↓ 0":
                        cell_diff.fill = fill_yellow

    wb.save(result_file)
    return FileResponse(result_file, filename="Processed_Output.xlsx")
